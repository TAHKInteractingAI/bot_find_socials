# from openpyxl import load_workbook
# from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# from bs4 import BeautifulSoup
# from selenium.webdriver.common.by import By


# # Mở tệp excel
# workbook = load_workbook('new.xlsx')
# input_sheet = workbook.active

# # Tìm kiếm và cập nhật thông tin CEO
# driver = webdriver.Chrome()  # Hoặc trình duyệt bạn muốn sử dụng

# row_index = 3
# for row in input_sheet.iter_rows(min_row=3, values_only=True):
#     company_name = row[0] if row[0] else ""

#     query = f"{company_name} CEO" "Linkedin"
#     query = query.replace(" ", "%20")

#     search_url = f"https://www.google.com/search?q={query}"

#     driver.get(search_url)
#     driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
#     page_source = driver.page_source

#     soup = BeautifulSoup(page_source, "html.parser")
#     search_results = soup.find_all('a', href=True)

#     linkedin_links = []

#     for result in search_results:
#         link = result['href']
#         if 'linkedin.com' in link and '/status/' not in link and 'translate' not in link and 'login' not in link and 'search' not in link and 'company' not in link:
#             linkedin_links.append(link)

#     # Lưu tên CEO vào cột thứ 11
#     if linkedin_links:
#         ceo_link = linkedin_links[0]  # Lấy liên kết đầu tiên
#         ceo_name = ceo_link  # Giả sử rằng tên CEO là phần sau cùng của liên kết
#         input_sheet.cell(row=row_index, column=11, value=ceo_name)
#         row_index += 1

# # Lưu tệp Excel
# workbook.save('new.xlsx')

# # Đóng trình duyệt
# driver.quit()

from tkinter import Tk, Label, Button, filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By

# Hàm bắt đầu tìm kiếm
def start_search():
    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    workbook = load_workbook(filename)
    input_sheet = workbook.active

    # Tìm kiếm và cập nhật thông tin CEO
    driver = webdriver.Chrome()  # Hoặc trình duyệt bạn muốn sử dụng

    row_index = 3
    for row in input_sheet.iter_rows(min_row=3, values_only=True):
        company_name = row[0] if row[0] else ""

        query = f"{company_name} CEO" "Linkedin"
        query = query.replace(" ", "%20")

        search_url = f"https://www.google.com/search?q={query}"

        driver.get(search_url)
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
        page_source = driver.page_source

        soup = BeautifulSoup(page_source, "html.parser")
        search_results = soup.find_all('a', href=True)

        linkedin_links = []

        for result in search_results:
            link = result['href']
            if 'linkedin.com' in link and '/status/' not in link and 'translate' not in link and 'login' not in link and 'search' not in link and 'company' not in link:
                linkedin_links.append(link)

        # Lưu tên CEO vào cột thứ 11
        if linkedin_links:
            ceo_link = linkedin_links[0]  # Lấy liên kết đầu tiên
            ceo_name = ceo_link  # Giả sử rằng tên CEO là phần sau cùng của liên kết
            input_sheet.cell(row=row_index, column=11, value=ceo_name)
            row_index += 1

    # Lưu tệp Excel
    workbook.save(filename)

    # Đóng trình duyệt
    driver.quit()

# Tạo giao diện người dùng
root = Tk()
root.geometry('400x200')
root.title("Tool")

label = Label(root, text="Click below to select Excel file")
label.pack(pady=10)

button = Button(root, text="Start", command=start_search)
button.pack(pady=10)

root.mainloop()
