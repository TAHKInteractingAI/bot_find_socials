from tkinter import Tk, Label, Button, filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

# Hàm để đăng nhập trang Rocketreach
def login(driver, email, password):
    driver.get("https://rocketreach.co/login")
    time.sleep(5)

    email_field = driver.find_element(By.XPATH, '/html/body/div[1]/div[7]/div[2]/div/form/div[1]/fieldset/div/div[1]/div/label/input')
    email_field.send_keys(email)

    password_field = driver.find_element(By.XPATH, '/html/body/div[1]/div[7]/div[2]/div/form/div[1]/fieldset/div/div[2]/div/label/input')
    password_field.send_keys(password)

    login_button = driver.find_element(By.XPATH, '/html/body/div[1]/div[7]/div[2]/div/form/div[1]/fieldset/button')
    login_button.click()

    time.sleep(5)

# Hàm để lấy thông tin từ trang Rocketreach
def get_info(link, driver, email, password):
    login(driver, email, password)

    driver.get("https://rocketreach.co/person")
    time.sleep(5)

    search_box = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[2]/div[2]/div/div[2]/div[2]/div/div[1]/div/rr-keyword-search-facet-input/form/div/div/input')
    search_box.send_keys(link)

    search_button = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[2]/div[2]/div/div[2]/div[2]/div/div[1]/div/rr-keyword-search-facet-input/form/div/div/button/span')
    search_button.click()

    time.sleep(5)

    try:
        email_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/rr-unified-search-results/div/div[3]/div/ul/li[1]/div[1]/rr-profile-directive/div/div[1]/div[6]/div[1]/div/div[1]/ul')
        email = email_element.get_attribute('href').replace('mailto:', '')
        print("Email:", email)
    except:
        print("Email not found.")

    time.sleep(2)

# Hàm bắt đầu tìm kiếm
def start_search():
    driver = webdriver.Chrome()
    email = "dtrinhb3@gmail.com"  # Thay thế bằng địa chỉ email thực của bạn
    password = "Trinh2001@"  # Thay thế bằng mật khẩu thực của bạn

    filename = filedialog.askopenfilename(initialdir="/", title="Select file", filetypes=(("xlsx files", "*.xlsx"), ("all files", "*.*")))
    workbook = load_workbook(filename)
    input_sheet = workbook.active

    for row in input_sheet.iter_rows(min_row=1, values_only=True):
        link = row[0]
        get_info(link, driver, email, password)

    driver.quit()

# Tạo giao diện người dùng
root = Tk()
root.geometry('400x200')
root.title("Rocketreach Information Finder")

label = Label(root, text="Click below to select Excel file")
label.pack(pady=10)

button = Button(root, text="Select file", command=start_search)
button.pack(pady=10)

root.mainloop()
