import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import openpyxl
import tkinter as tk
from tkinter import filedialog
import re
import json
import time
import g4f

g4f.debug.logging = True  # Enable logging
g4f.check_version = False  # Disable automatic version checking
import re

from g4f.Provider import (
    AItianhu,
    Aichat,
    Bard,
    Bing,
    ChatBase,
    ChatgptAi,
    OpenaiChat,
    Vercel,
    You,
    Yqcloud,
)
import spacy


def clean_input(input_str):
    # Loại bỏ các ký tự đặc biệt, chỉ giữ lại ký tự chữ cái và dấu cách
    cleaned_str = re.sub(r"[^a-zA-Z\s]", "", input_str)
    return cleaned_str


def select_file():
    file_path = filedialog.askopenfilename(filetypes=(("All files", "*"),))
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file_path)


def search_twitter_profile():
    input_file_path = file_entry.get()
    twitter_data = []

    try:
        # Export the data to a new Excel file
        input_workbook = load_workbook(input_file_path, read_only=True)
        input_sheet = input_workbook.active
        driver = webdriver.Chrome()
        driver.get("https://www.google.com")
        time.sleep(5)
        lan = driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div/a[1]")
        if lan.text == "English":
            driver.get(lan.get_attribute("href"))
        # Create a list to store data in JSON format
        json_data = []
        i = 0

        for row in input_sheet.iter_rows(min_row=3, values_only=True):
            i = i + 1
            print("{}".format(i))
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""
                ceo_name = ""
                while not ceo_name:
                    query = f"{keywords} of {company_name} "
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    text = soup.get_text()
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    text = soup.get_text(strip=True)
                    org = f"The {keywords} of ABCDXYX is Metraclac Kousier"
                    response = g4f.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        provider=g4f.Provider.You,
                        messages=[
                            {
                                "role": "user",
                                "content": f"I will give you a some text about a {keywords} of a company.This is the text :BEGIN: {text[:500]} :END. According the given text, your answer should be just the real {keywords} name,  your answer format answer will be :{org}, no more explain. ",
                            }
                        ],
                        stream=True,
                    )

                    for message in response:
                        ceo_name += message
                    # Using regular expression to extract the name after "is"
                    ceo_name = "".join(ceo_name).replace("*", "")
                    match = re.search(r'is\s+([^"]+)', ceo_name)
                    if match:
                        ceo_name = match.group(1)
                    ceo_name = ceo_name[:-1] if ceo_name.endswith(".") else ceo_name
                    nlp = spacy.load("en_core_web_sm")
                    # Process the text with SpaCy
                    doc = nlp(ceo_name)
                    # Extract person names
                    ceo_name = [ent.text for ent in doc.ents if ent.label_ == "PERSON"]

                ceo_name = "".join(ceo_name).strip()

                if "not" in ceo_name:
                    ceo_name = f"{keywords} not found"
                    json_data.append(
                        {
                            "Name": ceo_name,
                            "Company Name": company_name,
                            "Keywords": keywords,
                            "Twitter Profile": "No Twitter links found.",
                        }
                    )
                else:
                    query = f"Twitter {ceo_name} {keywords} of {company_name}"
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
                    twitter_links = []
                    name_twitter = []
                    search_results = driver.find_elements(
                        By.CSS_SELECTOR, "div.v7W49e > div"
                    )
                    time.sleep(5)
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )

                    # Get the HTML content of the element
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    element_ = soup.find_all("div", jscontroller="SC7lYd")
                    # Extract the text or HTML of the element
                    if element_:
                        for element in element_:
                            # Or get the HTML of the element
                            html_inside_element = (
                                element.prettify()
                            )  # This retains the HTML structure
                            soup = BeautifulSoup(html_inside_element, "html.parser")
                            a_tag = soup.find("a")  # Find the <a> tag
                            if a_tag is not None:
                                link = a_tag.get("href")

                            h3_tag = soup.find("h3")  # Find the <a> tag
                            if h3_tag is not None:
                                name = h3_tag.get_text()
                                # Get the value of the href attribute
                                # link = soup.find_all("a", href=True)[0]
                                # print(len(link))
                                # name = soup.find_all("h3").text
                                # print(len(name))

                                if (
                                    "twitter.com" in link
                                    and "/status/" not in link
                                    and "/post/" not in link
                                    and "/posts/" not in link
                                    and "post" not in link
                                    and "posts" not in link
                                    and "story" not in link
                                    and "news" not in link
                                    and "job" not in link
                                    and "today" not in link
                                    and "author" not in link
                                    and "pulse" not in link
                                    and "company" not in link
                                    and "text" not in link
                                    and "translate" not in link
                                    and "login" not in link
                                    and "search" not in link
                                    and any(word in name for word in ceo_name.split())
                                ):
                                    print(link, name)
                                    twitter_links.append(link)
                                    ceo_name = name.split("-")[0].strip()
                                    name_twitter.append(name)
                                    break

                    if (
                        not twitter_links
                        or "posts" in twitter_links[-1]
                        or "/posts/" in twitter_links[-1]
                    ):
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "Twitter Profile": "No Twitter links found.",
                            }
                        )
                    else:
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "Twitter Profile": twitter_links[-1],
                            }
                        )
        driver.quit()

        json_file_path = "twitter_data.json"
        with open(json_file_path, "w") as json_file:
            json.dump(json_data, json_file, indent=2)

        # Save JSON data to Excel
        excel_output_file_path = "output_twitter.xlsx"
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(
            [
                "Name",
                "Company Name",
                "Keywords",
                "Twitter Profile",
            ]
        )

        for entry in json_data:
            output_sheet.append(
                [
                    entry["Name"],
                    entry["Company Name"],
                    entry["Keywords"],
                    entry["Twitter Profile"],
                ]
            )

        output_workbook.save(excel_output_file_path)

        status_label.config(
            text=f"Output saved to: {excel_output_file_path} and {json_file_path}"
        )

    except Exception as e:
        error_message = f"Error: {str(e)}"
        status_label.config(text=error_message)


def search_facebook_profile():
    input_file_path = file_entry.get()
    facebook_data = []

    try:
        # Export the data to a new Excel file
        input_workbook = load_workbook(input_file_path, read_only=True)
        input_sheet = input_workbook.active
        driver = webdriver.Chrome()
        driver.get("https://www.google.com")
        time.sleep(5)
        lan = driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div/a[1]")
        if lan.text == "English":
            driver.get(lan.get_attribute("href"))
        # Create a list to store data in JSON format
        json_data = []
        i = 0

        for row in input_sheet.iter_rows(min_row=3, values_only=True):
            i = i + 1
            print("{}".format(i))
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""
                ceo_name = ""
                while not ceo_name:
                    query = f"{keywords} of {company_name} "
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    text = soup.get_text()
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    text = soup.get_text(strip=True)
                    org = f"The {keywords} of ABCDXYX is Metraclac Kousier"
                    response = g4f.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        provider=g4f.Provider.You,
                        messages=[
                            {
                                "role": "user",
                                "content": f"I will give you a some text about a {keywords} of a company.This is the text :BEGIN: {text[:500]} :END. According the given text, your answer should be just the real {keywords} name,  your answer format answer will be :{org}, no more explain. ",
                            }
                        ],
                        stream=True,
                    )

                    for message in response:
                        ceo_name += message
                    # Using regular expression to extract the name after "is"
                    ceo_name = "".join(ceo_name).replace("*", "")
                    match = re.search(r'is\s+([^"]+)', ceo_name)
                    if match:
                        ceo_name = match.group(1)
                    ceo_name = ceo_name[:-1] if ceo_name.endswith(".") else ceo_name
                    nlp = spacy.load("en_core_web_sm")
                    # Process the text with SpaCy
                    doc = nlp(ceo_name)
                    # Extract person names
                    ceo_name = [ent.text for ent in doc.ents if ent.label_ == "PERSON"]

                ceo_name = "".join(ceo_name).strip()

                if "not" in ceo_name:
                    ceo_name = f"{keywords} not found"
                    json_data.append(
                        {
                            "Name": ceo_name,
                            "Company Name": company_name,
                            "Keywords": keywords,
                            "Facebook Profile": "No facebook links found.",
                        }
                    )
                else:
                    query = f"Facebook {ceo_name} {keywords} of {company_name}"
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
                    facebook_links = []
                    name_facebook = []
                    search_results = driver.find_elements(
                        By.CSS_SELECTOR, "div.v7W49e > div"
                    )
                    time.sleep(5)
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )

                    # Get the HTML content of the element
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    element_ = soup.find_all("div", jscontroller="SC7lYd")
                    # Extract the text or HTML of the element
                    if element_:
                        for element in element_:
                            # Or get the HTML of the element
                            html_inside_element = (
                                element.prettify()
                            )  # This retains the HTML structure
                            soup = BeautifulSoup(html_inside_element, "html.parser")
                            a_tag = soup.find("a")  # Find the <a> tag
                            if a_tag is not None:
                                link = a_tag.get("href")

                            h3_tag = soup.find("h3")  # Find the <a> tag
                            if h3_tag is not None:
                                name = h3_tag.get_text()
                                # Get the value of the href attribute
                                # link = soup.find_all("a", href=True)[0]
                                # print(len(link))
                                # name = soup.find_all("h3").text
                                # print(len(name))

                                if (
                                    "facebook.com" in link
                                    and "/status/" not in link
                                    and "/post/" not in link
                                    and "/posts/" not in link
                                    and "post" not in link
                                    and "posts" not in link
                                    and "story" not in link
                                    and "news" not in link
                                    and "job" not in link
                                    and "today" not in link
                                    and "author" not in link
                                    and "pulse" not in link
                                    and "company" not in link
                                    and "text" not in link
                                    and "translate" not in link
                                    and "login" not in link
                                    and "search" not in link
                                    and any(word in name for word in ceo_name.split())
                                ):
                                    print(link, name)
                                    facebook_links.append(link)
                                    ceo_name = name.split("-")[0].strip()
                                    name_facebook.append(name)
                                    break

                    if (
                        not facebook_links
                        or "posts" in facebook_links[-1]
                        or "/posts/" in facebook_links[-1]
                    ):
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "Facebook Profile": "No Facebook links found.",
                            }
                        )
                    else:
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "Facebook Profile": facebook_links[-1],
                            }
                        )
        driver.quit()

        json_file_path = "facebook_data.json"
        with open(json_file_path, "w") as json_file:
            json.dump(json_data, json_file, indent=2)

        # Save JSON data to Excel
        excel_output_file_path = "output_facebook.xlsx"
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(
            [
                "Name",
                "Company Name",
                "Keywords",
                "Facebook Profile",
            ]
        )

        for entry in json_data:
            output_sheet.append(
                [
                    entry["Name"],
                    entry["Company Name"],
                    entry["Keywords"],
                    entry["Facebook Profile"],
                ]
            )

        output_workbook.save(excel_output_file_path)

        status_label.config(
            text=f"Output saved to: {excel_output_file_path} and {json_file_path}"
        )

    except Exception as e:
        error_message = f"Error: {str(e)}"
        status_label.config(text=error_message)


def search_linkedin_profile():
    input_file_path = file_entry.get()
    linkedin_data = []

    try:
        # Export the data to a new Excel file
        input_workbook = load_workbook(input_file_path, read_only=True)
        input_sheet = input_workbook.active
        driver = webdriver.Chrome()
        driver.get("https://www.google.com")
        time.sleep(5)
        lan = driver.find_element(By.XPATH, "/html/body/div[1]/div[4]/div/div/a[1]")
        if lan.text == "English":
            driver.get(lan.get_attribute("href"))
        # Create a list to store data in JSON format
        json_data = []
        i = 0

        for row in input_sheet.iter_rows(min_row=3, values_only=True):
            i = i + 1
            print("{}".format(i))
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""
                ceo_name = ""
                while not ceo_name:
                    query = f"{keywords} of {company_name} "
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    text = soup.get_text()
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    text = soup.get_text(strip=True)
                    org = f"The {keywords} of ABCDXYX is Metraclac Kousier"
                    response = g4f.ChatCompletion.create(
                        model="gpt-3.5-turbo",
                        provider=g4f.Provider.You,
                        messages=[
                            {
                                "role": "user",
                                "content": f"I will give you a some text about a {keywords} of a company.This is the text :BEGIN: {text[:500]} :END. According the given text, your answer should be just the real {keywords} name,  your answer format answer will be :{org}, no more explain. ",
                            }
                        ],
                        stream=True,
                    )

                    for message in response:
                        ceo_name += message
                    # Using regular expression to extract the name after "is"
                    ceo_name = "".join(ceo_name).replace("*", "")
                    match = re.search(r'is\s+([^"]+)', ceo_name)
                    if match:
                        ceo_name = match.group(1)
                    ceo_name = ceo_name[:-1] if ceo_name.endswith(".") else ceo_name
                    nlp = spacy.load("en_core_web_sm")
                    # Process the text with SpaCy
                    doc = nlp(ceo_name)
                    # Extract person names
                    ceo_name = [ent.text for ent in doc.ents if ent.label_ == "PERSON"]

                ceo_name = "".join(ceo_name).strip()

                if "not" in ceo_name:
                    ceo_name = f"{keywords} not found"
                    json_data.append(
                        {
                            "Name": ceo_name,
                            "Company Name": company_name,
                            "Keywords": keywords,
                            "LinkedIn Profile": "No LinkedIn links found.",
                        }
                    )
                else:
                    query = f"Linkedin {ceo_name} {keywords} of {company_name}"
                    query = query.replace(" ", "%20")
                    search_url = f"https://www.google.com/search?q={query}"
                    driver.get(search_url)
                    time.sleep(5)
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
                    linkedin_links = []
                    name_linkedin = []
                    search_results = driver.find_elements(
                        By.CSS_SELECTOR, "div.v7W49e > div"
                    )
                    time.sleep(5)
                    element = driver.find_element(
                        By.CSS_SELECTOR, "div.eqAnXb > div#search"
                    )

                    # Get the HTML content of the element
                    element_html = element.get_attribute("outerHTML")
                    soup = BeautifulSoup(element_html, "html.parser")
                    element_ = soup.find_all("div", jscontroller="SC7lYd")
                    # Extract the text or HTML of the element
                    if element_:
                        for element in element_:
                            # Or get the HTML of the element
                            html_inside_element = (
                                element.prettify()
                            )  # This retains the HTML structure
                            soup = BeautifulSoup(html_inside_element, "html.parser")
                            a_tag = soup.find("a")  # Find the <a> tag
                            if a_tag is not None:
                                link = a_tag.get("href")

                            h3_tag = soup.find("h3")  # Find the <a> tag
                            if h3_tag is not None:
                                name = h3_tag.get_text()
                                # Get the value of the href attribute
                                # link = soup.find_all("a", href=True)[0]
                                # print(len(link))
                                # name = soup.find_all("h3").text
                                # print(len(name))

                                if (
                                    "linkedin.com" in link
                                    and "/status/" not in link
                                    and "/post/" not in link
                                    and "/posts/" not in link
                                    and "post" not in link
                                    and "posts" not in link
                                    and "story" not in link
                                    and "news" not in link
                                    and "job" not in link
                                    and "today" not in link
                                    and "author" not in link
                                    and "pulse" not in link
                                    and "company" not in link
                                    and "text" not in link
                                    and "translate" not in link
                                    and "login" not in link
                                    and "search" not in link
                                    and any(word in name for word in ceo_name.split())
                                ):
                                    print(link, name)
                                    linkedin_links.append(link)
                                    ceo_name = name.split("-")[0].strip()
                                    name_linkedin.append(name)
                                    break

                    if (
                        not linkedin_links
                        or "posts" in linkedin_links[-1]
                        or "/posts/" in linkedin_links[-1]
                    ):
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "LinkedIn Profile": "No LinkedIn links found.",
                            }
                        )
                    else:
                        json_data.append(
                            {
                                "Name": ceo_name,
                                "Company Name": company_name,
                                "Keywords": keywords,
                                "LinkedIn Profile": linkedin_links[-1],
                            }
                        )
        driver.quit()

        json_file_path = "linkedin_data.json"
        with open(json_file_path, "w") as json_file:
            json.dump(json_data, json_file, indent=2)

        # Save JSON data to Excel
        excel_output_file_path = "output_linkedin.xlsx"
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(
            [
                "Name",
                "Company Name",
                "Keywords",
                "LinkedIn Profile",
            ]
        )

        for entry in json_data:
            output_sheet.append(
                [
                    entry["Name"],
                    entry["Company Name"],
                    entry["Keywords"],
                    entry["LinkedIn Profile"],
                ]
            )

        output_workbook.save(excel_output_file_path)

        status_label.config(
            text=f"Output saved to: {excel_output_file_path} and {json_file_path}"
        )

    except Exception as e:
        error_message = f"Error: {str(e)}"
        status_label.config(text=error_message)


root = tk.Tk()
root.title("Find Social Media account")

# create the main frame
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack()

# create the file frame
file_frame = tk.LabelFrame(main_frame, text="Attached input file")
file_frame.pack(fill="x", padx=10, pady=10)

file_entry = tk.Entry(file_frame, width=40)
file_entry.pack(side="left", padx=10, pady=5)

file_button = tk.Button(file_frame, text="Browser", command=select_file)
file_button.pack(side="left", padx=10, pady=5)
# create the status label
status_label = tk.Label(main_frame, text="", font=("Arial", 12))
status_label.pack(pady=10)

# create the send button
send_button = tk.Button(
    main_frame, text="Find Twitter Account", command=search_twitter_profile
)
send_button.pack(pady=5)

# create the send button
send_button = tk.Button(
    main_frame, text="Find Facebook Account", command=search_facebook_profile
)
send_button.pack(pady=5)

# create the send button
send_button = tk.Button(
    main_frame, text="Find LinkedIn Account", command=search_linkedin_profile
)
send_button.pack(pady=5)

root.mainloop()
